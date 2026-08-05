import io

import openpyxl
import pandas as pd
import streamlit as st

# Configuración de la página
st.set_page_config(
    page_title="Procesador de Productos | Palermo",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# Estilos visuales — no afectan la lógica del procesamiento
st.markdown(
    """
    <style>
        @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=Playfair+Display:wght@600;700&display=swap');

        .stApp {
            background:
                radial-gradient(circle at 6% 5%, rgba(190, 161, 110, 0.18), transparent 24%),
                radial-gradient(circle at 94% 0%, rgba(48, 57, 45, 0.10), transparent 25%),
                #f7f6f2;
            color: #1f2520;
            font-family: "DM Sans", sans-serif;
        }

        .block-container {
            max-width: 1160px;
            padding-top: 2.4rem;
            padding-bottom: 3rem;
        }

        .hero {
            background: linear-gradient(135deg, #202820 0%, #344332 100%);
            color: #ffffff;
            padding: 2.6rem 2.8rem;
            border-radius: 22px;
            margin-bottom: 1.5rem;
            box-shadow: 0 14px 35px rgba(31, 43, 31, 0.18);
        }

        .hero-kicker {
            color: #d8bd81;
            font-size: 0.76rem;
            font-weight: 700;
            letter-spacing: 0.14em;
            text-transform: uppercase;
            margin-bottom: 0.7rem;
        }

        .hero h1 {
            font-family: "Playfair Display", serif;
            font-size: clamp(2rem, 4vw, 3rem);
            line-height: 1.12;
            margin: 0 0 0.7rem 0;
            color: #ffffff;
        }

        .hero p {
            margin: 0;
            color: #e5e9e2;
            font-size: 1rem;
            line-height: 1.6;
            max-width: 720px;
        }

        .steps-container {
            display: flex;
            gap: 0.8rem;
            flex-wrap: wrap;
            margin: 0.25rem 0 1.4rem 0;
        }

        .step {
            flex: 1;
            min-width: 200px;
            display: flex;
            align-items: center;
            gap: 0.75rem;
            background: rgba(255, 255, 255, 0.74);
            border: 1px solid #e5e3dc;
            border-radius: 14px;
            padding: 0.85rem 1rem;
        }

        .step-number {
            display: flex;
            align-items: center;
            justify-content: center;
            min-width: 30px;
            height: 30px;
            border-radius: 50%;
            background: #d8bd81;
            color: #263025;
            font-size: 0.82rem;
            font-weight: 700;
        }

        .step-text {
            color: #4c554a;
            font-size: 0.88rem;
            line-height: 1.3;
        }

        .step-text strong {
            color: #263025;
            display: block;
            font-size: 0.92rem;
        }

        .upload-card, .preview-card {
            background: rgba(255, 255, 255, 0.86);
            border: 1px solid #e6e3da;
            border-radius: 18px;
            padding: 1.7rem;
            box-shadow: 0 8px 24px rgba(36, 44, 34, 0.05);
        }

        .card-label {
            color: #9c7b3e;
            font-size: 0.74rem;
            font-weight: 700;
            letter-spacing: 0.12em;
            text-transform: uppercase;
            margin-bottom: 0.35rem;
        }

        .card-title {
            color: #263025;
            font-family: "Playfair Display", serif;
            font-size: 1.45rem;
            margin: 0 0 0.35rem 0;
        }

        .card-description {
            color: #687064;
            margin: 0 0 1.15rem 0;
            font-size: 0.94rem;
        }

        .metric-card {
            background: #263025;
            border-radius: 15px;
            padding: 1.1rem 1.25rem;
            color: white;
            margin-bottom: 1rem;
        }

        .metric-card .metric-label {
            color: #d7ddcf;
            font-size: 0.76rem;
            font-weight: 600;
            letter-spacing: 0.08em;
            text-transform: uppercase;
        }

        .metric-card .metric-value {
            color: #e5c987;
            font-family: "Playfair Display", serif;
            font-size: 2rem;
            font-weight: 700;
            line-height: 1.15;
            margin-top: 0.25rem;
        }

        [data-testid="stFileUploader"] {
            border: 1.5px dashed #b6a26e;
            border-radius: 13px;
            padding: 0.5rem;
            background: #fcfbf8;
        }

        [data-testid="stFileUploader"] section {
            padding: 1.1rem;
        }

        .stDownloadButton > button {
            width: 100%;
            border: 0;
            border-radius: 10px;
            background: #b79554;
            color: #1f2820;
            font-weight: 700;
            padding: 0.7rem 1rem;
            transition: all 0.2s ease;
        }

        .stDownloadButton > button:hover {
            background: #d2b774;
            color: #172016;
            transform: translateY(-1px);
        }

        [data-testid="stDataFrame"] {
            border: 1px solid #ebe8df;
            border-radius: 12px;
            overflow: hidden;
        }

        @media (max-width: 700px) {
            .block-container {
                padding-top: 1.2rem;
                padding-left: 1rem;
                padding-right: 1rem;
            }

            .hero {
                padding: 1.8rem 1.4rem;
            }

            .upload-card, .preview-card {
                padding: 1.2rem;
            }
        }
    </style>
    """,
    unsafe_allow_html=True,
)

# Encabezado
st.markdown(
    """
    <section class="hero">
        <div class="hero-kicker">Gestión de catálogo</div>
        <h1>Procesador de Productos Nuevos</h1>
        <p>
            Cargá el archivo de Palermo, detectá los productos destacados en amarillo
            y descargá el Excel final listo para usar.
        </p>
    </section>
    """,
    unsafe_allow_html=True,
)

# Indicador visual de proceso
st.markdown(
    """
    <div class="steps-container">
        <div class="step">
            <div class="step-number">1</div>
            <div class="step-text"><strong>Subí el Excel</strong>Seleccioná el archivo inicial.</div>
        </div>
        <div class="step">
            <div class="step-number">2</div>
            <div class="step-text"><strong>Procesamiento automático</strong>Se detectan los productos nuevos.</div>
        </div>
        <div class="step">
            <div class="step-number">3</div>
            <div class="step-text"><strong>Descargá el resultado</strong>Obtené el archivo listo para cargar.</div>
        </div>
    </div>
    """,
    unsafe_allow_html=True,
)

# Área de carga
st.markdown(
    """
    <section class="upload-card">
        <div class="card-label">Archivo de origen</div>
        <h2 class="card-title">Cargá tu planilla de productos</h2>
        <p class="card-description">
            Se aceptan archivos Excel en formato <strong>.xlsx</strong> o <strong>.xls</strong>.
            Los productos nuevos deben estar identificados con color amarillo en la columna A.
        </p>
    </section>
    """,
    unsafe_allow_html=True,
)

uploaded_file = st.file_uploader(
    "Seleccioná el archivo Excel inicial (por ejemplo, “Carga Palermo”)",
    type=["xlsx", "xls"],
    label_visibility="visible",
)


def es_fila_encabezado(ws, r):
    """Verifica si la fila actual es un encabezado de bloque (ej. contiene FECHA o COSTO)."""
    val_b = str(ws.cell(row=r, column=2).value or "").upper()
    val_c = str(ws.cell(row=r, column=3).value or "").upper()
    return "FECHA" in val_b or "COSTO" in val_c or "TARJETA" in val_c


def procesar_excel(file_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    registro